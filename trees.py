import re
import time

import nltk
import sqlite3
import csv
import openpyxl

from text_processing.preprocess import preprocess
from text_processing.translate import translate
from text_processing.dates import process_dates


class Corpus:

    def __init__(self, db, table):

        self.db = db
        self.table = table
        self.conn = sqlite3.connect(f"db/{db}.db")
        self.conn.row_factory = sqlite3.Row
        self.c = self.conn.cursor()
        self.c.execute("SELECT * FROM " + table)
        self.topics = []
        self.data = []

        raw_data = self.c.fetchall()

        for row in raw_data:
            doc = Document(row, self.conn, table)
            doc.process(['title','lead','content'])
            self.data.append(doc)

    def find_topics(self):

        for row1 in self.data:
            rows_except_this = [r for r in self.data if r.url != row1.url and r.country != row1.country]
            for row2 in rows_except_this:
                com_words = row1.named_entities['content'].intersection(row2.named_entities['content'])
                if len(com_words) and com_words not in self.topics:
                    self.topics.append(com_words)

        self.topics.sort(key=lambda s: -len(s))


class Document:

    def __init__(self, row, conn, table):

        types = ('title', 'lead', 'content')
        self.raw = row

        self.country = row['country']
        self.date = row['date']
        self.orig_data = dict.fromkeys(types)
        self.orig_data['title'] = row['title']

        try:
            self.orig_data['lead'] = row['lead']
        except IndexError:
            self.orig_data['lead'] = row['description']

        self.orig_data['content'] = row['content']
        self.url = row['reference']

        self.translated = dict.fromkeys(types)
        self.double_translated = dict.fromkeys(types)
        self.tokens = dict.fromkeys(types)
        self.named_entities = dict.fromkeys(types)

        self.conn = conn
        self.table = table

        self.dates = process_dates(list(self.tokens)).append(self.date)

    def process(self, arr_of_types):

        for typ in arr_of_types:

            if typ == 'content':
                col = 'translated'
                col1 = 'translated1'

            else:
                col = f'translated_{typ}'
                col1 = f'translated1_{typ}'

            if self.raw[col]:
                self.translated[typ] = self.raw[col]
            else:
                self.double_translate(typ)

            if self.raw[col1]:
                self.double_translated[typ] = self.raw[col1]
            else:
                self.double_translate(typ)

            self.tokens[typ] = [word for word in preprocess(self.translated[typ]) if
                                word in preprocess(self.double_translated[typ])]

            parse_tree = nltk.ne_chunk(nltk.tag.pos_tag(self.tokens[typ]),
                                       binary=True)  # POS tagging before chunking!

            self.named_entities[typ] = {k[0] for branch in parse_tree.subtrees() for k in list(branch) if branch.label() == 'NE'}
            self.unite_countries_in(typ,'nes')
            self.find_entities(typ)
            self.unite_countries_in(typ,'nes')

    def find_entities(self, ty):

        text = re.findall(r"[\w]+|[^\s\w]", self.translated[ty])

        to_remove = set()
        to_add = set()

        nes = self.named_entities[ty]

        for ent1 in nes:
            if ent1 in text:
                idx1 = text.index(ent1)
                entities_except_this = nes - set(ent1)

                for ent2 in entities_except_this:
                    if ent2 in text:
                        idx2 = text.index(ent2)
                        if ((idx2 - idx1 == 2) and (text[idx1+1] == ' ' or text[idx1+1] == '-'
                                                    or text[idx1+1] == "'" or text[idx1+1] == 'of')) or idx2 - idx1 == 1:

                            united_entity = ' '.join([ent1, ent2])
                            to_add.add(united_entity)
                            to_remove.add(ent1)
                            to_remove.add(ent2)

        self.named_entities[ty] = (self.named_entities[ty] - to_remove) | to_add

    def unite_countries_in(self, ty, type_of_data):
        conn = sqlite3.connect("db/countries.db")
        c = conn.cursor()
        c.execute("SELECT * FROM countries")
        all_rows = c.fetchall()
        to_remove = set()
        to_add = set()
        if type_of_data == "nes":
            data = self.named_entities[ty]
        elif type_of_data=='tokens':
            data = self.tokens[ty]
        for ent in data:
                for row in all_rows:
                    low = [w.lower() for w in row if w is not None]
                    if ent.lower() in low:
                        to_remove.add(ent)
                        to_add.add(row[0])

        if type_of_data == 'nes':
            self.named_entities[ty] = (self.named_entities[ty] - to_remove) | to_add
        elif type_of_data == 'tokens':
            self.tokens[ty].extend(list(to_add))
            self.tokens[ty] = [t for t in self.tokens[ty] if t not in to_remove]

    def double_translate(self, ty):

        n = 1500  # length limit
        text = self.orig_data[ty]
        self.translated[ty] = ''
        self.double_translated[ty] = ''

        if "Краткое описание: " in text:
            text = text.split("Краткое описание: ")[1]

        # Split into parts of 1500 before translating
        text = [text[i:i + n] for i in range(0, len(text), n)]

        for part in text:

            eng_text = translate(part)
            orig_text = translate(eng_text, self.country)
            eng1_text = translate(orig_text)

            self.translated[ty] += ' '
            self.translated[ty] += eng_text
            self.double_translated[ty] += ' '
            self.double_translated[ty] += eng1_text

        c = self.conn.cursor()
        if ty == 'content':
            col = 'translated'
            col1 = 'translated1'

        else:
            col = f'translated_{ty}'
            col1 = f'translated1_{ty}'

        c.execute(f"UPDATE {self.table} SET {col}=(?), {col1}=(?) WHERE reference=(?)",
                  (self.translated[ty], self.double_translated[ty], self.url))
        self.conn.commit()


class Node:

    def __init__(self, item):

        self.name = item
        self.level = len(self.name)

        self.parents = []
        self.children = []
        self.documents = []
        self.percents = []
        self.subtopics = []

    def has_free_links(self, other_node):

        topic = self.name
        my_free_words = [d.named_entities['content'] - topic for d in self.documents]
        suit_documents = [sd for sd in self.documents if sd not in other_node.documents]
        par_free_words = [pd.named_entities['content'] - topic for pd in suit_documents]
        connection = 0

        for doc in my_free_words:
            for d in par_free_words:
                if doc.intersection(d):
                    connection += 1

        if my_free_words and par_free_words:
            percent_of_connected = connection/(len(my_free_words)*len(par_free_words))
            return percent_of_connected
        return 0

    def find_most_frequent(self, type):
        freq_words = {}
        for document in self.documents:

            # descr = []
            # descr.extend(document.tokens['title'])
            # descr.extend(document.tokens['lead'])lead
            descr = set(document.tokens['title']) | set(document.tokens['lead'])

            if type=='lowercase':
                parse_tree = nltk.ne_chunk(nltk.tag.pos_tag(descr),
                                           binary=True)  # POS tagging before chunking!

                nes = {k[0] for branch in parse_tree.subtrees() for k in list(branch) if
                                            branch.label() == 'NE'}
                descr = [d for d in descr if d not in nes]

            for word in descr:

                if word in freq_words.keys():
                    freq_words[word] +=1
                else:
                    freq_words[word] = 1
        if freq_words:
            maxval = max(freq_words.values())
            result = {key:value for key,value in freq_words.items() if value==maxval}
            return result
        else:
            return freq_words

    def add_document(self, doc):
        self.documents.append(doc)

    def assign_parent(self, parent):
        if parent not in self.parents:
            self.parents.append(parent)
            self.percents.append(self.has_free_links(parent))

    def assign_child(self, child):
        if child not in self.children:
            self.children.append(child)

    def isparent(self, other_node):
        if other_node in self.parents:
            return True
        return False

    def ischild(self, other_node):
        if other_node in self.children:
            return True
        return False

    def isroot(self):
        if self.parents:
            return False
        return True

    def isleaf(self):
        if self.children:
            return False
        return True

    def contains(self, children):
        for ch in children:
            if self.name.issubset(ch.name):
                return True
        return False


class Tree:

    def __init__(self, corpus):

        self.nodes = [Node(item) for item in corpus.topics]
        self.assign_documents(corpus)
        self.db = corpus.db

        for node in self.nodes:

            for other_node in [n for n in self.nodes if n != node]:

                if node.name.issubset(other_node.name) and node not in other_node.children and not node.contains(other_node.children):
                    node.assign_parent(other_node)
                    other_node.assign_child(node)

                if other_node.name.issubset(node.name) and node not in other_node.parents and not other_node.contains(node.children):
                    other_node.assign_parent(node)
                    node.assign_child(other_node)

        self.roots = [n for n in self.nodes if n.isroot()]
        write_topics_to_xl("корневые темы.xlsx", self.roots)
        self.last_nodes = set()

    def assign_documents(self, corpus):

        for node in self.nodes:
            for doc in corpus.data:
                if node.name.issubset(doc.named_entities['content']):
                    node.add_document(doc)

    def find_last_topics(self):
        for root in self.roots:
            x = check_free([root])
            if x:
                self.last_nodes.add(x)
        write_topics_to_xl("после отсекания.xlsx", self.last_nodes)


    # def unite_similar_topics(self):

        # to_remove = self.delete_subsets()
        #
        # self.last_nodes -= to_remove
        # write_topics_to_xl("удалили дубликаты.xlsx", self.last_nodes)
        # to_remove = set()

        # Это проверка

        # for node in self.last_nodes:
        #     percent = 0
        #     for doc in node.documents:
        #         nes = doc.title_named_entities|doc.lead_named_entities
        #         if nes.intersection(node.name):
        #             percent += 1
        #     percent /= len(node.documents)
        #     if percent < 0.5:
        #         to_remove.add(node)
        #
        # self.last_nodes -= to_remove
        # write_topics_to_xl("правило 50%.xlsx", self.last_nodes)
        # to_remove = set()
        #
        # for n in self.last_nodes:
        #     frequent_words_dict = self.find_most_frequent(n)
        #     if any(k < 0.5*len(n.documents) for k in frequent_words_dict.values()):
        #         to_remove.add(n)
        #     else:
        #         rest = []
        #
        #         for d in n.documents:
        #             for w in frequent_words_dict.values():
        #                 if w not in d.tokens['title'] and w not in d.tokens['lead']:
        #                     rest.append(d)
        #                     continue
        #
        #         has_these_words = [d for d in n.documents if d not in rest]
        #
        #         for doc in rest:
        #             perc = 0
        #             for doc1 in has_these_words:
        #                 if set(doc.tokens['title']).intersection(set(doc1.tokens['title'])):
        #                     perc +=1
        #             perc /= len(has_these_words)
        #             if perc < 0.5:
        #                 n.documents.remove(doc)
        #
        # self.last_nodes -= to_remove
        #
        # write_topics_to_xl("проверка.xlsx", self.last_nodes, with_children=True)
        #
        # tr_1, ta_1 = self.add_and_remove(principle="country")
        # self.last_nodes = (self.last_nodes - tr_1) | ta_1
        # to_remove = self.delete_subsets()
        # self.last_nodes -= to_remove
        #
        # write_topics_to_xl("после объединения.xlsx", self.last_nodes, with_children=True)

        # nodes = sorted(list(self.last_nodes), key=lambda n: len(n.name))
        # old_nodes = self.last_nodes
        #
        # while old_nodes:
        #     to_remove = set()
        #     to_add = set()
        #
        #     for node in self.last_nodes:
        #         other_nodes = self.last_nodes-{node}
        #
        #         for other_node in other_nodes:
        #             if all(elem in node.documents for elem in other_node.documents):  #node contains all of other_node
        #                 new_node = Node(node.name | other_node.name)
        #                 new_node.subtopics = [node, other_node]
        #                 new_node.documents = node.documents
        #                 old_nodes.remove(node)
        #                 old_nodes.remove(other_node)
        #                 # to_remove.add(node)
        #                 # to_remove.add(other_node)
        #                 # to_add.add(new_node)
        #             if node.name.issubset(other_node.name):
        #                 # to_remove.add(node)
        #                 # to_remove.add(other_node)
        #                 old_nodes.remove(node)
        #                 old_nodes.remove(other_node)
        #                 new_node = Node(node.name & other_node.name)
        #                 new_node.documents = []
        #                 new_node.documents.extend(node.documents)
        #                 new_node.documents.extend([doc for doc in other_node.documents if doc not in new_node.documents])
        #                 to_add.add(new_node)
        #
        #     self.last_nodes = (self.last_nodes - to_remove) | to_add
        #     print(len(old_nodes))
        #
        # write_topics_to_xl("3.csv", self.last_nodes, with_children=True)

    def delete_subsets(self):
        to_remove = set()

        for n in self.last_nodes:
            other_nodes = self.last_nodes - {n}
            for on in other_nodes:
                if n.name.issubset(on.name) and on not in to_remove:
                    to_remove.add(n)

        return to_remove

    def add_and_remove(self, principle):
        all_links = []
        to_remove = set()
        to_add = set()

        for node in self.last_nodes:
            countries = {d.country for d in node.documents}
            print(node.name)
            print(countries)
            nodes_except_this = self.last_nodes - {node}

            for other_node in nodes_except_this:

                # if principle == "documents":
                #     common_documents = set(node.documents).intersection(set(other_node.documents))
                #     percent_1 = len(common_documents) / len(node.documents)
                #     percent_2 = len(common_documents) / len(other_node.documents)
                #     if percent_1 >= 0.5 and percent_2 >= 0.5:
                #         print(node.name, other_node.name)
                #         for cd in common_documents:
                #             print(translate(cd.title))
                #         print(percent_1, percent_2)
                #         all_links[node][other_node] = percent_1
                # elif principle == 'country':

                other_countries = {d.country for d in other_node.documents}

                common_documents = set(node.documents).intersection(set(other_node.documents))
                common_countries = {d.country for d in common_documents}
                percent_1 = len(common_countries) / len(countries)
                percent_2 = len(common_countries) / len(other_countries)
                # if common_countries:
                #     print(node.name, other_node.name)
                #     print(countries)
                #     print(other_countries)
                #     print(percent_1)
                #     print(percent_2)
                #     for cd in common_documents:
                #         print(cd.title)

                # other_news = dict.fromkeys(other_countries, set())
                # for k in other_news.keys():
                #         for doc in node.documents:
                #             if doc.country == k:
                #                 other_news[k].add(doc)
                #         if k in news.keys():
                #             print(node.name)
                #             print(other_node.name)
                #             l = [doc.title for doc in news[k] if doc in other_news[k]]
                #             print(l)
                #             if l:
                #
                #                 percent_1 += 1
                #                 percent_2 += 1
                #                 continue

                # percent_1 /= len(news.keys())
                # percent_2 /= len(other_news.keys())
                if percent_1 > 0.5 and percent_2 > 0.5:
                        all_links.append((node, other_node, percent_1*percent_2))

        for node in self.last_nodes:
            try:
                l = []
                for a in all_links:
                    if a[0] == node:
                        l.append((a[1], a[2]))

                max_value = max(l, key=lambda x: x[1])[1]
                max_sim_nodes = [s[0] for s in l if s[1] == max_value]
                name = node.name
                subtopics = []
                documents = node.documents
                to_remove.add(node)
                print(f"Node: {node.name}")

                for msn in max_sim_nodes:
                    print(f"Max_sim_node: {msn.name}")
                    name |= msn.name
                    subtopics.append(msn)

                    documents.extend(msn.documents)

                    to_remove.add(msn)

                new_node = Node(name)
                new_node.subtopics = subtopics
                new_node.documents = list(set(documents))
                to_add.add(new_node)

                # new_node.children = [node, maximum_similarity_node]
            except ValueError:
                continue

        return to_remove, to_add


def check_free(nodes):
    for node in nodes:
        for ch in node.children:
            ch_percent = ch.percents[ch.parents.index(node)]
            for gch in ch.children:
                gchild_percent = gch.percents[gch.parents.index(ch)]
                if gchild_percent < ch_percent and gchild_percent < 0.5:
                    return ch
                elif gch.children:
                    return check_free(ch.children)
                else:
                    return gch


def check_xl(nodes, all_links):
    file = csv.writer(open("t.csv", "w"), delimiter=',')

    for n in nodes:
        countries = {d.country for d in n.documents}
        row = [f"{n.name} | {countries}"]
        for k, m in all_links[n].items():
            row.append(f"{k.name} | {m} | {' '.join({d.country for d in k.documents})}")
            cd = set(n.documents).intersection(set(k.documents))
            cc = {d.country for d in cd}
            perc = len(cc)/len(countries)
            row.append(perc)
            row.append(cc)
            tex = ' | '.join([d.title for d in cd])
            row.append(tex)
        file.writerow(row)


def write_topics_to_xl(fname, nodes, with_children=False):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Topic'
    sheet['B1'] = 'News'
    sheet['C1'] = 'Keywords'
    for i, n in enumerate(nodes):
        if with_children:
            text = ''
            if n.subtopics:
                for c in n.subtopics:
                    text += ' {'.join(c.name)
                    text += ' | '
            else:
                text = ' '.join(n.name)
        else:
            text = ' '.join(n.name)
        sheet.cell(row=i+2, column=1).value = text
        docs = n.documents
        for j,doc in enumerate(docs):
            sheet.cell(row=i+2, column=j+2).value = f"{doc.country} | {doc.translated['title']} | {doc.translated['lead']} | {doc.url} | {doc.translated['content']} | {doc.named_entities['title']} | {doc.named_entities['content']}"
    wb.save(fname)


def write_words_to_xl(fname, data):
    file = csv.writer(open(fname, "w"), delimiter=',')
    headers = ['Document', 'Deleted words']
    file.writerow(headers)

    for r in data:
        row = [r.translated, r.removed_words]
        file.writerow(row)


def write_start_words_to_xl(fname, data):
    file = csv.writer(open(fname, "w"))
    headers = ['Document', 'Deleted words']
    file.writerow(headers)

    for r in data:
        for key, ind in r.start_words.items():
            row = [key, ind]
            file.writerow(row)


if __name__ == '__main__':

    now = time.time()

    db = input("DB name: ")
    table = input("Table name: ")

    if not db:
        db = "day"
    if not table:
        table = "buffer"

    c = Corpus(db, table)
    c.find_topics()
    # write_words_to_xl("double_translation.csv", c.data)
    # write_start_words_to_xl("start_words.csv", c.data)

    t = Tree(c)
    t.find_last_topics()
    # t.unite_similar_topics()
    # write_topics_to_xl(f"{db}-topics.csv", t.last_nodes)
    # f = open("time1.txt", "w")
    # f.write(str(time.time()-now))