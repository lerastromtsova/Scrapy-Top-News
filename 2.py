""" С маленькими буквами, хороший вариант - итоговый"""


import openpyxl
from trees import Node, Document
from text_processing.preprocess import preprocess

def write_topics_to_xl(fname, nodes):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'Topic'
    sheet['B1'] = 'News'
    sheet['C1'] = 'Keywords'
    for i, n in enumerate(nodes):
        maxval = n.find_most_frequent("uppercase")
        t = ''
        for k, v in maxval.items():
                    t += f"| {k}:{v} |"

        text = ' '.join(n.name) + '|' +t
        sheet.cell(row=i+2, column=1).value = text
        docs = n.documents
        for j,doc in enumerate(docs):
            sheet.cell(row=i+2, column=j+2).value = f"{doc.country} | {doc.translated['title']} | {doc.translated['lead']} | {doc.url} | {doc.translated['content']} | {doc.named_entities['title']} | {doc.named_entities['content']}"
    wb.save(fname)

wb = openpyxl.load_workbook("после отсекания.xlsx")
sheet = wb.active

nodes = set()

for j,row in enumerate(sheet.iter_rows()):
    if j != 0:
            name = row[0].value.split('|')
            node = Node(item=name)

            for i in range(1,len(row)):
                if row[i].value is not None:
                    doc = row[i].value.split("|")
                    country = doc[0]
                    title = doc[1]
                    lead = doc[2]
                    roww = {'country': country,
                           'title': title,
                           'lead': lead,
                           'date': '',
                           'content': '',
                           'reference': '',
                           }
                    d = Document(roww, '', '')
                    d.tokens = {'title': preprocess(title), 'lead': preprocess(lead), 'content': ''}
                    d.translated = {'title': title, 'lead': lead, 'content': ''}
                    # d.unite_countries_in('title', 'tokens')
                    # d.unite_countries_in('lead', 'tokens')
                    node.documents.append(d)

            nodes.add(node)

to_remove = set()

for n in nodes:
            frequent_words_dict = n.find_most_frequent("uppercase")

            if any(k < 0.5*len(n.documents) for k in frequent_words_dict.values()):
                to_remove.add(n)
            else:
                rest = []

                for d in n.documents:
                    for w in frequent_words_dict.keys():
                        if w not in d.tokens['title'] and w not in d.tokens['lead']:
                            rest.append(d)
                            continue

                has_these_words = [d for d in n.documents if d not in rest]

                for doc in rest:
                    perc = 0
                    for doc1 in has_these_words:
                        if set(doc.tokens['title']).intersection(set(doc1.tokens['title'])):
                            perc +=1
                    if has_these_words:
                        perc /= len(has_these_words)
                        if perc < 0.5:
                            if doc in n.documents:
                                n.documents.remove(doc)
                    else:
                        if doc in n.documents:
                            n.documents.remove(doc)

            countries = {d.country for d in n.documents}
            if len(countries) == 1:
                to_remove.add(n)

nodes -= to_remove
write_topics_to_xl("проверка.xlsx", nodes)


most_similar = []
while nodes:
    n = nodes.pop()
    other_nodes = nodes.copy()
    freq_words_1 = set(n.find_most_frequent("lowercase").keys())
    similar = [n]
    while other_nodes:
        on = other_nodes.pop()

        freq_words_2 = set(on.find_most_frequent("lowercase").keys())
        percent1 = 0
        percent2 = 0
        for doc in on.documents:
            text = []
            text.extend(doc.tokens['title'])
            text.extend(doc.tokens['lead'])
            if any([word in text for word in freq_words_1]):
                percent1 += 1
        percent1 /= len(on.documents)
        for doc in n.documents:
            text = []
            text.extend(doc.tokens['title'])
            text.extend(doc.tokens['lead'])
            if any([word in text for word in freq_words_2]):
                percent2 += 1
        percent2 /= len(n.documents)
        if percent1 > 0.5 and percent2>0.5:
            nodes.remove(on)
            similar.append(on)
    most_similar.append(similar)

for ms in most_similar:
    for s in ms:
        print(s.name)
        print(s.find_most_frequent('lowercase'))

    print("|")




# tr_1, ta_1 = self.add_and_remove(principle="country")
# self.last_nodes = (self.last_nodes - tr_1) | ta_1
# to_remove = self.delete_subsets()
# self.last_nodes -= to_remove
#
# write_topics_to_xl("после объединения.xlsx", self.last_nodes, with_children=True)