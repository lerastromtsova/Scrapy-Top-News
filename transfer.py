import openpyxl
import sqlite3

STOP_WORDS_FOR_STATES = {"Southern",
                         "Western",
                         "Eastern",
                         "Northern",
                         "Central"}


def transfer_states(xl_file, db_name, table):
    workbook = openpyxl.load_workbook(xl_file, read_only=True)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)
    states = []

    conn = sqlite3.connect(f"db/{db_name}.db")
    c = conn.cursor()

    for row in worksheet.iter_rows():
        cell = row[7].value

        if cell not in STOP_WORDS_FOR_STATES:
            counts = states.count(cell)
            if counts == 5:
                c.execute("INSERT INTO " + table + "(name) VALUES (?)", (cell,))

            states.append(cell)
    conn.commit()
    conn.close()


transfer_states("db/worldcities.xlsx", "countries", "state")

def transfer_cities(xl_file, db_name, table):
    workbook = openpyxl.load_workbook(xl_file, read_only=True)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    conn = sqlite3.connect(f"db/{db_name}.db")
    c = conn.cursor()

    for row in worksheet.iter_rows():
        city_name = row[0].value
        city_name_ascii = row[1].value
        state = row[7].value
        c.execute("SELECT id FROM state WHERE name=(?)",(state,))
        try:
            i = c.fetchone()[0]
            c.execute("INSERT INTO " + table + "(name, name_ascii, state_id) VALUES (?,?,?)", (city_name,city_name_ascii,i))
        except TypeError:
            continue

    conn.commit()
    conn.close()


# transfer_cities("db/worldcities.xlsx", "countries", "cities")

